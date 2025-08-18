from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file
from flask_login import login_required, current_user
from collections import defaultdict
from io import BytesIO
import calendar
from datetime import datetime, date

from .models import User, Mission, Assignment, OnCallAssignment
from .forms import UserForm, EditUserForm, MissionForm, AssignmentForm, OnCallAssignmentForm
from . import db

import openpyxl
from openpyxl.utils import get_column_letter

main = Blueprint('main', __name__)

def get_month_dates(year, month):
    _, num_days = calendar.monthrange(year, month)
    return [date(year, month, day) for day in range(1, num_days+1)]

@main.route('/')
@login_required
def dashboard():
    now = datetime.now()
    year = request.args.get('year', now.year, type=int)
    month = request.args.get('month', now.month, type=int)
    filter_user = request.args.get('user_id', type=int)
    filter_mission = request.args.get('mission_id', type=int)

    users = User.query.order_by(User.username).all()
    missions = Mission.query.order_by(Mission.name).all()

    if filter_user:
        users = [u for u in users if u.id == filter_user]
    if filter_mission:
        missions = [m for m in missions if m.id == filter_mission]

    dates = get_month_dates(year, month)
    first_day = date(year, month, 1)
    last_day = date(year, month, calendar.monthrange(year, month)[1])

    assignments = Assignment.query.filter(Assignment.date >= first_day, Assignment.date <= last_day).all()
    on_calls = OnCallAssignment.query.filter(OnCallAssignment.date >= first_day, OnCallAssignment.date <= last_day).all()

    if filter_user:
        assignments = [a for a in assignments if a.user_id == filter_user]
        on_calls = [o for o in on_calls if o.user_id == filter_user]
    if filter_mission:
        assignments = [a for a in assignments if a.mission_id == filter_mission]
        on_calls = [o for o in on_calls if o.mission_id == filter_mission]

    assignment_map = {d: {m.id: [] for m in missions} for d in dates}
    on_call_map = {d: {m.id: [] for m in missions} for d in dates}

    for a in assignments:
        if a.mission_id in [m.id for m in missions]:
            assignment_map[a.date][a.mission_id].append(a.user.username)
    for oc in on_calls:
        if oc.mission_id in [m.id for m in missions]:
            on_call_map[oc.date][oc.mission_id].append(oc.user.username)

    user_summary = defaultdict(lambda: defaultdict(int))
    for a in assignments:
        user_summary[a.user.username][a.mission.name] += 1

    return render_template('dashboard.html',
                           year=year, month=month, dates=dates,
                           users=users, missions=missions,
                           assignment_map=assignment_map, on_call_map=on_call_map,
                           user_summary=user_summary,
                           all_users=User.query.order_by(User.username).all(),
                           all_missions=Mission.query.order_by(Mission.name).all(),
                           filter_user=filter_user, filter_mission=filter_mission)

@main.route('/users', methods=['GET', 'POST'])
@login_required
def users():
    if current_user.role != 'admin':
        flash('Access denied.')
        return redirect(url_for('main.dashboard'))
    form = UserForm()
    if form.validate_on_submit():
        if User.query.filter_by(username=form.username.data).first():
            flash('User already exists.')
        else:
            u = User(username=form.username.data, role=form.role.data)
            u.set_password(form.password.data)
            db.session.add(u)
            db.session.commit()
            flash('User added.')
            return redirect(url_for('main.users'))
    user_list = User.query.order_by(User.username).all()
    return render_template('users.html', form=form, users=user_list)

@main.route('/users/edit/<int:user_id>', methods=['GET', 'POST'])
@login_required
def edit_user(user_id):
    if current_user.role != 'admin':
        flash('Access denied.')
        return redirect(url_for('main.dashboard'))
    user = User.query.get_or_404(user_id)
    from .forms import EditUserForm
    form = EditUserForm(obj=user)
    if form.validate_on_submit():
        user.username = form.username.data
        user.role = form.role.data
        if form.password.data:
            user.set_password(form.password.data)
        db.session.commit()
        flash('User updated.')
        return redirect(url_for('main.users'))
    return render_template('edit_user.html', form=form, user=user)

@main.route('/users/delete/<int:user_id>', methods=['POST'])
@login_required
def delete_user(user_id):
    if current_user.role != 'admin':
        flash('Access denied.')
        return redirect(url_for('main.dashboard'))
    user = User.query.get_or_404(user_id)
    if user.username == 'admin':
        flash('Cannot delete default admin.')
    else:
        db.session.delete(user)
        db.session.commit()
        flash('User deleted.')
    return redirect(url_for('main.users'))

@main.route('/missions', methods=['GET', 'POST'])
@login_required
def missions():
    if current_user.role != 'admin':
        flash('Access denied.')
        return redirect(url_for('main.dashboard'))
    form = MissionForm()
    if form.validate_on_submit():
        if Mission.query.filter_by(name=form.name.data).first():
            flash('Mission already exists.')
        else:
            db.session.add(Mission(name=form.name.data))
            db.session.commit()
            flash('Mission added.')
            return redirect(url_for('main.missions'))
    missions = Mission.query.order_by(Mission.name).all()
    return render_template('missions.html', form=form, missions=missions)

@main.route('/assign', methods=['GET', 'POST'])
@login_required
def assign():
    if current_user.role != 'admin':
        flash('Access denied.')
        return redirect(url_for('main.dashboard'))
    form = AssignmentForm()
    if form.validate_on_submit():
        exists = Assignment.query.filter_by(user_id=form.user_id.data, mission_id=form.mission_id.data, date=form.date.data).first()
        if exists:
            flash('Assignment already exists.')
        else:
            db.session.add(Assignment(user_id=form.user_id.data, mission_id=form.mission_id.data, date=form.date.data))
            db.session.commit()
            flash('Assignment created.')
            return redirect(url_for('main.dashboard'))
    return render_template('assign.html', form=form, title="Assign Mission")

@main.route('/oncall_assign', methods=['GET', 'POST'])
@login_required
def oncall_assign():
    if current_user.role != 'admin':
        flash('Access denied.')
        return redirect(url_for('main.dashboard'))
    form = OnCallAssignmentForm()
    if form.validate_on_submit():
        exists = OnCallAssignment.query.filter_by(user_id=form.user_id.data, mission_id=form.mission_id.data, date=form.date.data).first()
        if exists:
            flash('On-call assignment already exists.')
        else:
            db.session.add(OnCallAssignment(user_id=form.user_id.data, mission_id=form.mission_id.data, date=form.date.data))
            db.session.commit()
            flash('On-call assignment created.')
            return redirect(url_for('main.dashboard'))
    return render_template('assign.html', form=form, title="Assign On-Call")

@main.route('/export_excel')
@login_required
def export_excel():
    now = datetime.now()
    year = request.args.get('year', now.year, type=int)
    month = request.args.get('month', now.month, type=int)
    dates = get_month_dates(year, month)
    first_day = date(year, month, 1)
    last_day = date(year, month, calendar.monthrange(year, month)[1])

    assignments = Assignment.query.filter(Assignment.date >= first_day, Assignment.date <= last_day).all()
    on_calls = OnCallAssignment.query.filter(OnCallAssignment.date >= first_day, OnCallAssignment.date <= last_day).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year}-{month:02d}"

    header = ["User", "Mission", "Date", "Type"]
    ws.append(header)

    for a in assignments:
        ws.append([a.user.username, a.mission.name, a.date.strftime('%Y-%m-%d'), "Assigned"])
    for oc in on_calls:
        ws.append([oc.user.username, oc.mission.name, oc.date.strftime('%Y-%m-%d'), "On-Call"])

    # Autosize columns
    for col in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(12, length + 2)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    from flask import send_file
    return send_file(bio, as_attachment=True,
                     download_name=f"mission_summary_{year}_{month:02d}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
